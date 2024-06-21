import openpyxl
import sys

def populate_pre_gaps(filename):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(filename)
    
    # Select the sheet named 'concat_data'
    sheet = wb['concat']
    
    # Extract data from columns 'Inning' and 'Player'
    innings = [cell.value for cell in sheet['J'] if cell.value is not None][1:]
    first_inning = next((value for value in innings if value is not None), None)


    players = [cell.value for cell in sheet['K'] if cell.value is not None][1:]
    first_player = next((value for value in players if value is not None), None)

    num_lines = sheet.max_row - 1

    if(num_lines > len(innings)):
    
        gap = num_lines - len(innings)

        new_innings = [first_inning] * gap + innings
        new_players = [first_player] * gap + players

        # Delete existing columns 'J' and 'K'
        sheet.delete_cols(10, 2)  # 'J' is the 10th column, 'K' is the 11th column
        
        # Insert new columns 'new_innings' and 'new_players'
        sheet.insert_cols(10, 2)  # Insert 2 columns starting at the 10th position
        for i, value in enumerate(new_innings, start=2):
            sheet.cell(row=i, column=10).value = value
        for i, value in enumerate(new_players, start=2):
            sheet.cell(row=i, column=11).value = value

        sheet.cell(row=1, column=10).value = "Player"
        sheet.cell(row=1, column=11).value = "Inning"

        
        # Save the workbook
        wb.save(filename)
        print("Now all gaps are filled!")

    else:
        print("there is no gaps!")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <filename.xlsx>")
        sys.exit(1)
    
    filename = sys.argv[1]
    populate_pre_gaps(filename)

