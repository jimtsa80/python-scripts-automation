import sys
import warnings
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def update_excel_file(file_path):
    # Ignore warnings
    warnings.simplefilter(action='ignore', category=FutureWarning)
    warnings.simplefilter(action='ignore', category=UserWarning)


    print(f"Loading Excel file: {file_path}")

    try:
        # Load the workbook using openpyxl
        workbook = load_workbook(file_path)
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        sys.exit(1)

    try:
        # Convert 'Batters' sheet to DataFrame
        batters_data = workbook['Batters']
        batters_df = pd.DataFrame(batters_data.values)
        batters_df.columns = batters_df.iloc[0]  # Set the first row as the header
        batters_df = batters_df[1:]  # Skip the first row

        # Convert 'new' sheet to DataFrame
        new_data = workbook['new']
        new_df = pd.DataFrame(new_data.values)
        new_df.columns = new_df.iloc[0]  # Set the first row as the header
        new_df = new_df[1:]  # Skip the first row
    except KeyError as e:
        print(f"Error reading sheets: {e}")
        sys.exit(1)

    print("Selecting and renaming columns from 'Batters'")
    try:
        batters_selected = batters_df[['Location', 'Inning', 'Sequence Frame Number']]
        batters_selected = batters_selected.copy() 
        batters_selected.rename(columns={'Location': 'Player'}, inplace=True)
    except KeyError as e:
        print(f"Error selecting columns: {e}")
        sys.exit(1)

    print("Merging dataframes")
    try:
        merged_df = pd.merge(new_df, batters_selected, on='Sequence Frame Number', how='left')
    except KeyError as e:
        print(f"Error merging dataframes: {e}")
        sys.exit(1)

    print("Saving the updated Excel file")
    try:
        # Create a new sheet named 'concat' and append the merged data
        if 'concat' in workbook.sheetnames:
            workbook.remove(workbook['concat'])
        
        concat_sheet = workbook.create_sheet(title='concat')

        for r_idx, row in enumerate(dataframe_to_rows(merged_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                concat_sheet.cell(row=r_idx, column=c_idx, value=value)

        # Save the updated workbook
        workbook.save(file_path)
    except Exception as e:
        print(f"Error saving the Excel file: {e}")
        sys.exit(1)

    print(f"Updated file saved to {file_path}")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <path_to_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    update_excel_file(file_path)
