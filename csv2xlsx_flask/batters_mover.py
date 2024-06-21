import sys
import openpyxl

def copy_and_sort_excel(file_path):
    # Load the workbook
    workbook = openpyxl.load_workbook(file_path)
    
    # Load the sheets
    batters_sheet = workbook["Batters"]
    homeplate_sheet = workbook["Homeplate"]

    # Get existing data from Homeplate sheet
    existing_data = list(homeplate_sheet.values)
    
    # Get headers and data
    headers = existing_data[0]
    existing_data = existing_data[1:]  # Skip the headers

    # Copy data from Batters sheet, skipping the header
    new_data = list(batters_sheet.values)[1:]

    # Combine existing and new data
    combined_data = existing_data + new_data

    # Find the index of the "Sequence Frame Number" column
    sequence_frame_index = headers.index("Sequence Frame Number")

    # Sort the combined data by "Sequence Frame Number"
    sorted_data = sorted(combined_data, key=lambda x: x[sequence_frame_index])

    # Clear the Homeplate sheet
    homeplate_sheet.delete_rows(2, homeplate_sheet.max_row)

    # Re-add the headers if not already present
    if homeplate_sheet.max_row == 0:
        homeplate_sheet.append(headers)

    # Append the sorted data
    for row in sorted_data:
        homeplate_sheet.append(row)

    # Save the workbook
    workbook.save(file_path)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <path_to_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    copy_and_sort_excel(file_path)
