import sys
import pandas as pd
from openpyxl import load_workbook

def process_excel(file_path):
    print("Reading Excel file...")
    # Read the specified sheet from the Excel file
    df = pd.read_excel(file_path, sheet_name='concat')
    
    # Trim whitespace from string columns
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    
    # Define the columns to group by
    group_by_columns = ['Brand', 'Location', 'Time the brand is at screen', 'Screen Location', 'Screen Size %', 'Average Hits', 'Player', 'Inning']
    
    print("Grouping and aggregating data...")
    # Group by the specified columns and aggregate Duration and Total Hits
    grouped = df.groupby(group_by_columns).agg({
        'Duration': 'sum',
        'Total Hits': 'sum',
        'Sequence Frame Number': 'min'
    }).reset_index()
    
    # Sort by Sequence Frame Number
    grouped = grouped.sort_values(by='Sequence Frame Number')
    
    # Reorder columns to match the desired order
    ordered_columns = ['Brand', 'Location', 'Time the brand is at screen', 'Duration', 'Screen Location', 
                       'Screen Size %', 'Total Hits', 'Average Hits', 'Sequence Frame Number', 'Player', 'Inning']
    grouped = grouped[ordered_columns]
    
    # Rename the sheet to 'HomePlate'
    print("Renaming sheet to 'HomePlate'...")
    grouped_sheet_name = 'HomePlate'
    grouped.rename(columns={'Sequence Frame Number': 'First Sequence Frame Number'}, inplace=True)
    
    # Save the output to a new sheet in the same Excel file
    print("Saving to a new sheet...")
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        try:
            # Attempt to remove the existing 'HomePlate' sheet
            writer.book.remove(writer.book[grouped_sheet_name])
        except KeyError:
            pass  # If 'HomePlate' sheet doesn't exist, proceed
        grouped.to_excel(writer, sheet_name=grouped_sheet_name, index=False)
    
    # Remove unnecessary sheets using openpyxl
    print("Removing unnecessary sheets...")
    wb = load_workbook(file_path)
    sheets_to_remove = ['new', 'concat', 'Batters']
    for sheet_name in sheets_to_remove:
        if sheet_name in wb.sheetnames:
            wb.remove(wb[sheet_name])
            print(f"Sheet '{sheet_name}' removed successfully.")
        else:
            print(f"Sheet '{sheet_name}' not found, skipping...")
    wb.save(file_path)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <file_path>")
        sys.exit(1)
    
    file_path = sys.argv[1]
    process_excel(file_path)
