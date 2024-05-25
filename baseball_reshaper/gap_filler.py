import pandas as pd
import sys
from openpyxl import load_workbook

def fill_empty_cells(file_path):
    # Tab name to work with
    sheet_name = "concat"

    # Load the Excel file
    wb = load_workbook(file_path)

    # Get the existing worksheet
    ws = wb[sheet_name]

    # Load data into DataFrame
    df = pd.DataFrame(ws.values)
    df.columns = df.iloc[0]
    df = df.iloc[1:]

    # Forward fill the empty cells in the "Inning" and "Player" columns
    print("Filling empty cells...")
    df['Inning'] = df['Inning'].fillna(method='ffill')
    df['Player'] = df['Player'].fillna(method='ffill')

    # Remove completely empty rows
    df = df.dropna(how='all')

    # Clear existing data in the worksheet
    print("Clearing existing data in '{}' tab...".format(sheet_name))
    ws.delete_rows(2, ws.max_row)

    # Write the updated DataFrame back to the same Excel tab starting from row 2
    print("Writing updated data to '{}' tab starting from row 2...".format(sheet_name))
    for index, row in df.iterrows():
        ws.append(row.tolist())

    # Save the updated Excel file
    print("Saving changes to Excel file...")
    wb.save(file_path)

    print("Process completed.")

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py <file_path>")
        sys.exit(1)

    file_path = sys.argv[1]
    fill_empty_cells(file_path)

