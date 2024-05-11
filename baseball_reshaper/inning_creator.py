import openpyxl
import sys

def add_inning_column(input_file):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file)
    
    # Select the "Batters" worksheet
    ws = wb["Batters"]
    
    # Initialize variables for tracking innings
    current_brand = None
    inning = 1
    inning_half = 'top'
    
    # Iterate through rows and populate the "Inning" column
    for row_num in range(2, ws.max_row + 1):
        brand = ws.cell(row=row_num, column=1).value  # Assuming brand is in the first column
        if current_brand is None:
            current_brand = brand
        elif brand != current_brand:
            current_brand = brand
            inning = 1
            inning_half = 'top'
        else:
            if inning_half == 'top':
                inning += 1
                inning_half = 'bottom'
            else:
                inning_half = 'top'
        
        # Insert the "Inning" value into the row
        ws.cell(row=row_num, column=10, value=f"{inning_half} {inning}")  # Assuming Inning column is the tenth column

    # Save the modified workbook
    wb.save(input_file)

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script.py input_file.xlsx")
        sys.exit(1)

    input_file = sys.argv[1]
    add_inning_column(input_file)

